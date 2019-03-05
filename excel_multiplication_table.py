#! python3

"""
Create multiplication table in .xlsx format


"""

import openpyxl

wb = openpyxl.Workbook()

sheet = wb.get_active_sheet()

def createRows(maxNum, multi):
    """
    Creates base structure for the multiplier

    :param maxNum: Maximum Number for the multiplication table
    :param multi: The multiplication steps
    :return:
    """

    for i in range(1, maxNum+1, multi):
        sheet["A"+str(i+1)] = i #assigns numbers to vertically
        sheet.cell(row=1, column=(i+1)).value= i #assigns numbers horizontally

def multiply():
    mincell = "B2" #Cell for multiplication to start
    maxcell = sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate #Cell for multiplication to end

    table = sheet[mincell:maxcell]

    for row in table:
        for cell in row:
            cell_coord = cell.coordinate
            # Formula that you want to create = "A" column + cell_coord number x cell_coord letter + "1"
            x_multi = str(sheet[cell_coord].column)+"1"
            y_multi = "A"+str(sheet[cell_coord].row)

            multi_formula = "="+x_multi+"*"+y_multi #generate multiplication excel formula

            sheet[cell_coord] = multi_formula # insert formula into cells

if __name__ == "__main__":
    createRows(5,1)
    multiply()

wb.save("Multiplication.xlsx")
